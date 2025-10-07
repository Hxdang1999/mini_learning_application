# app/api/course_controller.py
from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from app.services.course_service import CourseService
from app.models.enrollment import Enrollment
from app.repositories.auth_repository import AuthRepository
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from app.models.course import Course
from app.models.assignment import Assignment
from app.models.user import User
from app.models.assignment import Submission
from app import db
from flask import send_file, jsonify

course_bp = Blueprint('course_bp', __name__, url_prefix='/api/courses')
course_service = CourseService()
auth_repo = AuthRepository()

@course_bp.route('/', methods=['POST'])
@jwt_required()
def create_course():
    current_user_info = get_jwt_identity()
    user_id = current_user_info['id']
    user_role = current_user_info['role']
    
    if user_role != 'teacher':
        return jsonify({"message": "Forbidden. Only teachers can create courses."}), 403
    data = request.get_json()
    title = data.get('title')
    description = data.get('description')
    is_public = data.get('is_public', False)
    response, status_code = course_service.create_course(user_id, title, description, is_public)
    return jsonify(response), status_code

@course_bp.route('/teacher', methods=['GET'])
@jwt_required()
def get_teacher_courses():
    try:
        current_user_info = get_jwt_identity()
        user_id = current_user_info['id']
        courses = course_service.get_teacher_courses(user_id)
        courses_list = [{
            "id": c.id,
            "title": c.title,
            "description": c.description,
            "is_public": c.is_public,
            "created_at": c.created_at.strftime('%Y-%m-%d %H:%M:%S') if c.created_at else 'N/A'
        } for c in courses]
        return jsonify({"courses": courses_list}), 200
    except Exception as e:
        return jsonify({"message": f"An error occurred: {str(e)}"}), 500

# app/api/course_controller.py
@course_bp.route('/<int:course_id>', methods=['GET'])
@jwt_required()
def get_course(course_id):
    current_user_info = get_jwt_identity()
    user_id = current_user_info['id']
    user_role = current_user_info['role']
    
    course = course_service.course_repo.get_by_id(course_id)
    if not course:
        return jsonify({"message": "Course not found"}), 404
    
    if user_role == 'teacher' and course.teacher_id != user_id:
        return jsonify({"message": "Forbidden. You are not the teacher of this course."}), 403
    
    if user_role == 'student':
        enrollment = course_service.course_repo.get_enrollment(user_id, course_id)
        if not enrollment or enrollment.status != 'active':
            return jsonify({"message": "Forbidden. You are not enrolled or your enrollment is not active."}), 403
    
    course_data = {
        "id": course.id,
        "title": course.title,
        "description": course.description,
        "is_public": course.is_public,
        "created_at": course.created_at.strftime('%Y-%m-%d %H:%M:%S') if course.created_at else 'N/A'
    }
    return jsonify(course_data), 200

@course_bp.route('/<int:course_id>', methods=['PUT'])
@jwt_required()
def update_course(course_id):
    current_user_info = get_jwt_identity()
    user_id = current_user_info['id']
    data = request.get_json()
    response, status_code = course_service.update_course(course_id, user_id, data)
    return jsonify(response), status_code

@course_bp.route('/<int:course_id>', methods=['DELETE'])
@jwt_required()
def delete_course(course_id):
    current_user_info = get_jwt_identity()
    user_id = current_user_info['id']
    response, status_code = course_service.delete_course(course_id, user_id)
    return jsonify(response), status_code

@course_bp.route('/available', methods=['GET'])
@jwt_required()
def get_available_courses():
    current_user_info = get_jwt_identity()
    user_role = current_user_info['role']
    student_id = current_user_info['id']
    if user_role != 'student':
        return jsonify({"message": "Forbidden. Only students can view available courses."}), 403
    courses = course_service.get_available_courses(student_id)
    courses_list = [{
        "id": c.id,
        "title": c.title,
        "description": c.description,
        "is_public": c.is_public
    } for c in courses]
    return jsonify({"courses": courses_list}), 200

@course_bp.route('/<int:course_id>/enroll', methods=['POST'])
@jwt_required()
def enroll_course(course_id):
    current_user_info = get_jwt_identity()
    user_id = current_user_info['id']
    user_role = current_user_info['role']
    if user_role != 'student':
        return jsonify({"message": "Forbidden. Only students can enroll in courses."}), 403
    response, status_code = course_service.enroll_student(user_id, course_id)
    return jsonify(response), status_code

@course_bp.route('/enrolled', methods=['GET'])
@jwt_required()
def get_enrolled_courses():
    current_user = get_jwt_identity()
    if current_user['role'] != 'student':
        return jsonify({"message": "Only students"}), 403
    enrollments = Enrollment.query.filter_by(student_id=current_user['id']).all()
    data = []
    for e in enrollments:
        data.append({
            "id": e.course.id,
            "title": e.course.title,
            "teacher_username": auth_repo.get_user_by_id(e.course.teacher_id).username,
            "is_public": e.course.is_public,
            "status": e.status
        })
    return jsonify({"courses": data}), 200

@course_bp.route('/<int:course_id>/unenroll', methods=['DELETE'])
@jwt_required()
def unenroll_course(course_id):
    current_user_info = get_jwt_identity()
    student_id = current_user_info['id']
    user_role = current_user_info['role']
    if user_role != 'student':
        return jsonify({"message": "Forbidden. Only students can unenroll from courses."}), 403
    response, status_code = course_service.unenroll_student(student_id, course_id)
    return jsonify(response), status_code

# Route mới: Lấy danh sách yêu cầu đăng ký pending
@course_bp.route('/<int:course_id>/pending', methods=['GET'])
@jwt_required()
def get_pending_enrollments(course_id):
    current_user = get_jwt_identity()
    if current_user['role'] != 'teacher':
        return jsonify({"message": "Only teachers can view pending enrollments"}), 403
    enrollments, status = course_service.get_pending_enrollments(current_user['id'], course_id)
    return jsonify(enrollments), status

# Route mới: Duyệt/từ chối yêu cầu đăng ký
@course_bp.route('/enrollments/<int:enrollment_id>/<string:action>', methods=['PUT'])
@jwt_required()
def handle_enrollment(enrollment_id, action):
    current_user = get_jwt_identity()
    if current_user['role'] != 'teacher':
        return jsonify({"message": "Only teachers can handle enrollments"}), 403
    approve = (action == "approve")
    response, status = course_service.handle_enrollment(current_user['id'], enrollment_id, approve)
    return jsonify(response), status

@course_bp.route('/<int:course_id>/lookup', methods=['GET'])
@jwt_required()
def lookup_course(course_id):
    current_user = get_jwt_identity()
    if current_user['role'] != 'student':
        return jsonify({"message": "Chỉ sinh viên mới có thể tra cứu khóa học"}), 403
    course = course_service.course_repo.get_by_id(course_id)
    if not course:
        return jsonify({"message": "Không tìm thấy khóa học"}), 404
    return jsonify({
        "id": course.id,
        "title": course.title,
        "description": course.description,
        "is_public": course.is_public
    }), 200

# Route mới: Lấy danh sách sinh viên active
@course_bp.route('/<int:course_id>/enrolled-students', methods=['GET'])
@jwt_required()
def get_enrolled_students(course_id):
    current_user = get_jwt_identity()
    if current_user['role'] != 'teacher':
        return jsonify({"message": "Only teachers can view enrolled students"}), 403
    enrollments, status = course_service.get_enrolled_students(current_user['id'], course_id)
    return jsonify(enrollments), status

# Route mới: Buộc rời sinh viên khỏi khóa học
@course_bp.route('/enrollments/<int:enrollment_id>/unenroll', methods=['DELETE'])
@jwt_required()
def teacher_unenroll(enrollment_id):
    current_user = get_jwt_identity()
    if current_user['role'] != 'teacher':
        return jsonify({"message": "Only teachers can unenroll students"}), 403
    response, status = course_service.teacher_unenroll_student(current_user['id'], enrollment_id)
    return jsonify(response), status

@course_bp.route('/<int:course_id>/export-grades', methods=['GET'])
@jwt_required()
def export_course_grades(course_id):
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from datetime import datetime
    from flask import send_file, jsonify

    current_user = get_jwt_identity()
    if current_user['role'] not in ['teacher', 'admin']:
        return jsonify({"message": "Permission denied"}), 403

    course = Course.query.get(course_id)
    if not course:
        return jsonify({"message": "Course not found"}), 404

    teacher = User.query.get(course.teacher_id)
    teacher_name = teacher.username if teacher else "—"
    assignments = Assignment.query.filter_by(course_id=course_id).all()
    enrollments = Enrollment.query.filter_by(course_id=course_id, status='active').all()
    students = [User.query.get(e.student_id) for e in enrollments if e.student_id]

    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng điểm"

    # ======= TIÊU ĐỀ =======
    ws.merge_cells("A1:E1")
    ws["A1"] = "Mini Learning Application"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append([f"Khóa học: {course.title} (ID: {course.id})"])
    ws.append([f"Giảng viên: {teacher_name} (ID: {teacher.id if teacher else '—'})"])
    ws.append([f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    ws.append([])

    # ======= HEADER =======
    headers = ["STT", "ID SV", "Tên sinh viên"]
    for a in assignments:
        headers.append(a.title)
    headers.append("Điểm trung bình")
    ws.append(headers)

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ======= DỮ LIỆU SINH VIÊN =======
    for idx, student in enumerate(students, start=1):
        row = [idx, student.id, student.username]
        scores = []
        for a in assignments:
            submission = Submission.query.filter_by(
                assignment_id=a.id, student_id=student.id
            ).first()
            if submission and submission.score is not None:
                score = float(submission.score)
                row.append(score)
                scores.append(score)
            else:
                row.append("—")
        avg = round(sum(scores) / len(scores), 2) if scores else "—"
        row.append(avg)
        ws.append(row)

    # ======= CĂN GIỮA KHÔNG TÔ MÀU =======
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ======= XUẤT FILE =======
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"course_{course_id}_grades.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
